import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def autofit_excel(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    wb.save(file_path)

def compare_excel_files(previous_file, current_file, output_file):
    df_previous = pd.read_excel(previous_file)
    df_this = pd.read_excel(current_file)

    # Ensure that compulsory columns are present
    if 'Main Code' not in df_previous.columns or 'Balance' not in df_previous.columns:
        raise ValueError("Previous file is missing required columns: 'Main Code' and 'Balance'")
    if 'Main Code' not in df_this.columns or 'Balance' not in df_this.columns:
        raise ValueError("Current file is missing required columns: 'Main Code' and 'Balance'")

    # Exclude rows with Limit == 0 if the 'Limit' column is present
    if 'Limit' in df_previous.columns:
        df_previous = df_previous[df_previous['Limit'] != 0]
    if 'Limit' in df_this.columns:
        df_this = df_this[df_this['Limit'] != 0]

    # Filter out rows where 'Main Code' is 'AcType Total' or 'Grand Total'
    df_previous = df_previous[~df_previous['Main Code'].isin(['AcType Total', 'Grand Total'])]
    df_this = df_this[~df_this['Main Code'].isin(['AcType Total', 'Grand Total'])]

    previous_codes = set(df_previous['Main Code'])
    this_codes = set(df_this['Main Code'])

    only_in_previous = df_previous.loc[df_previous['Main Code'].isin(previous_codes - this_codes)]
    only_in_this = df_this.loc[df_this['Main Code'].isin(this_codes - previous_codes)]
    in_both = df_previous.loc[df_previous['Main Code'].isin(previous_codes & this_codes)]

    # Safe merge and calculation of balance changes
    in_both = pd.merge(
        in_both[['Main Code', 'Balance']], 
        df_this[['Main Code', 'Balance']], 
        on='Main Code', 
        suffixes=('_previous', '_this')
    )
    in_both['Change'] = in_both['Balance_this'] - in_both['Balance_previous']

    opening_sum = df_previous['Balance'].sum()
    settled_sum = only_in_previous['Balance'].sum()
    new_sum = only_in_this['Balance'].sum()
    increase_decrease_sum = in_both['Change'].sum()
    adjusted_sum = opening_sum - settled_sum + new_sum + increase_decrease_sum
    closing_sum = df_this['Balance'].sum()

    opening_count = len(previous_codes)
    settled_count = len(previous_codes - this_codes)
    new_count = len(this_codes - previous_codes)
    closing_count = len(this_codes)

    reco_data = {
        'Description': ['Opening', 'Settled', 'New', 'Increase/Decrease', 'Adjusted', 'Closing'],
        'Amount': [opening_sum, settled_sum, new_sum, increase_decrease_sum, adjusted_sum, closing_sum],
        'No of Acs': [opening_count, settled_count, new_count, "", "", closing_count]
    }
    df_reco = pd.DataFrame(reco_data)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        only_in_previous.to_excel(writer, sheet_name='Settled', index=False)
        only_in_this.to_excel(writer, sheet_name='New', index=False)
        in_both.to_excel(writer, sheet_name='Movement', index=False)
        df_reco.to_excel(writer, sheet_name='Reco', index=False)

    autofit_excel(output_file)

    return output_file

def main():
    st.title("File Comparison Tool")

    st.write("Upload the previous period's Excel file and this period's Excel file to compare them. The Columns Required are Main Code and Balance. Get download link.")

    previous_file = st.file_uploader("Upload Previous Period's Excel File", type=["xlsx"])
    current_file = st.file_uploader("Upload This Period's Excel File", type=["xlsx"])

    if previous_file and current_file:
        output_file = 'comparison_output.xlsx'

        with st.spinner("Processing..."):
            with open('previous_file.xlsx', 'wb') as f:
                f.write(previous_file.getbuffer())

            with open('current_file.xlsx', 'wb') as f:
                f.write(current_file.getbuffer())

            result_file = compare_excel_files('previous_file.xlsx', 'current_file.xlsx', output_file)

        if result_file:
            with open(result_file, "rb") as file:
                st.download_button(
                    label="Download Comparison Output",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main()

